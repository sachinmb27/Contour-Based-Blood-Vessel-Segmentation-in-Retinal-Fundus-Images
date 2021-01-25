import cv2
import numpy as np
import os
from openpyxl import Workbook

loc = 'input/'
files = os.listdir(loc)
location = 'truth/'
files1 = os.listdir(location)
count1 = 0

#creating a new excel sheet to store the results
wb = Workbook()
ws = wb.active
ws['A1'] = 'Image'
ws['B1'] = 'True positive'
ws['C1'] = 'True negative'
ws['D1'] = 'False positive'
ws['E1'] = 'Flase negative'
ws['F1'] = 'Accuracy'
ws['G1'] = 'Sensitivity'
ws['H1'] = 'Specificity'

count = 2

for ff in files:
    file = loc + '/' + ff
    test_image = cv2.imread(file, 1)

    #test image is converted to LAB modal
    lab = cv2.cvtColor(test_image, cv2.COLOR_BGR2LAB)
    l, a, b = cv2.split(lab)

    #Contrast Limited Adaptive Histogram Equalization is applied
    clahe = cv2.createCLAHE(clipLimit=3.0)
    cl = clahe.apply(l)
    limg = cv2.merge((cl, a, b))

    #LAB modal converted back to RGB
    final = cv2.cvtColor(limg, cv2.COLOR_LAB2BGR)

    #applying alternate sequential filtering
    blue,green,red = cv2.split(final)
    r1 = cv2.morphologyEx(green, cv2.MORPH_OPEN, cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (5,5)), iterations = 1)
    R1 = cv2.morphologyEx(r1, cv2.MORPH_CLOSE, cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (5,5)), iterations = 1)
    r2 = cv2.morphologyEx(R1, cv2.MORPH_OPEN, cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (11,11)), iterations = 1)
    R2 = cv2.morphologyEx(r2, cv2.MORPH_CLOSE, cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (11,11)), iterations = 1)
    r3 = cv2.morphologyEx(R2, cv2.MORPH_OPEN, cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (23,23)), iterations = 1)
    R3 = cv2.morphologyEx(r3, cv2.MORPH_CLOSE, cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (23,23)), iterations = 1)
    f4 = cv2.subtract(R3, green)
    f5 = clahe.apply(f4)

    #tophat morphological transformation
    image1 = f5
    e_kernel = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (5,5))
    closeImg = cv2.morphologyEx(image1, cv2.MORPH_CLOSE, e_kernel)
    revImg = closeImg
    topHat = image1 - revImg

    #otsu with probability and minimization function
    imge = topHat
    blur = cv2.GaussianBlur(imge, (5,5), 0)
    hist = cv2.calcHist([blur], [0], None, [256], [0,256])
    hist_norm = hist.ravel()/hist.max()
    Q = hist_norm.cumsum()
    bins = np.arange(256)
    fn_min = np.inf
    thresh = -1
    for i in range(1, 256):
        p1, p2 = np.hsplit(hist_norm, [i]) #probabilities
        q1, q2 = Q[i],Q[255]-Q[i] #cum sum of classes
        b1, b2 = np.hsplit(bins,[i]) #weights

    #finding means and variances
    if q1 == 0:
        q1 = 0.0000001
    if q2 == 0:
        q2 = 0.0000001
    m1, m2 = np.sum(p1*b1)/q1, np.sum(p2*b2)/q2
    v1, v2 = np.sum(((b1-m1)**2)*p1)/q1,np.sum(((b2-m2)**2)*p2)/q2

    #calculates the minimization function
    fn = v1*q1 + v2*q2
    if fn < fn_min:
        fn_min = fn
        thresh = i

    #find otsu&'s threshold value with OpenCV function
    ret, otsu = cv2.threshold(blur, 0, 255, cv2.THRESH_BINARY+cv2.THRESH_OTSU)

    #removing very small contours through area parameter noise removal
    ret, f6 = cv2.threshold(f5, 15, 255, cv2.THRESH_BINARY)

    mask = np.ones(f5.shape[:2], dtype='uint8') * 255
    im2, contours, hierarchy = cv2.findContours(f6.copy(),cv2.RETR_TREE,cv2.CHAIN_APPROX_SIMPLE)
    for cnt in contours:
        if cv2.contourArea(cnt) <= 255:
            cv2.drawContours(mask, [cnt], -1, 0, -1)
    im = cv2.bitwise_and(f5, f5, mask=mask)
    ret, fin = cv2.threshold(im, 15, 255, cv2.THRESH_BINARY_INV)
    newfin = cv2.erode(fin, cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (3,3)), iterations=1)

    #removing blobs of unwanted size
    fundus_eroded = cv2.bitwise_not(newfin)
    xmask = np.ones(fundus_eroded.shape[:2], dtype='uint8') * 255
    x1, xcontours, xhierarchy = cv2.findContours(fundus_eroded.copy(), cv2.RETR_LIST, cv2.CHAIN_APPROX_SIMPLE)
    for cnt in xcontours:
        shape = 'unidentified'
        peri = cv2.arcLength(cnt, True)
        approx = cv2.approxPolyDP(cnt, 0.04 * peri, False)
        if len(approx) > 4 and cv2.contourArea(cnt) <= 3000 and cv2.contourArea(cnt) >= 100:
            shape = 'circle'
        else:
            shape = 'vessels'
        
        if(shape == 'circle'):
            cv2.drawContours(xmask, [cnt], -1, 0, -1)

    finimage = cv2.bitwise_and(fundus_eroded, fundus_eroded, mask=xmask)
    blood_vessels = cv2.bitwise_not(finimage)
    kernel = np.ones((2,2), np.uint8)
    blood_vessels = cv2.subtract(255, blood_vessels)

    #eroding is done to eliminate the minor noises (post processing)
    new = cv2.morphologyEx(blood_vessels, cv2.MORPH_OPEN, kernel)
    new1 = cv2.morphologyEx(new, cv2.MORPH_CLOSE, kernel)

    #saving the blood vessel extracted image
    cv2.imwrite('test/bv/' + ff + 'bvextracted.jpg', new1)

    #reading the masked image
    mask_image = cv2.imread('01_test_mask.jpg', 0)

    #reading the manually segmented blood vessel images for results calculation
    fil = location + '/' + files1[count1]
    count1 = count1 + 1
    truth_image = cv2.imread(fil, 0)
    prediction = new1

    for i in range(0, prediction.shape[0]):
        for j in range(0, prediction.shape[1]):
            if prediction[i, j] > 127:
                prediction[i, j] = 1
            else:
                prediction[i, j] = 0

    for i in range(0, truth_image.shape[0]):
        for j in range(0, truth_image.shape[1]):
            if truth_image[i, j] > 127:
                truth_image[i, j] = 1
            else:
                truth_image[i, j] = 0

    for i in range(0, mask_image.shape[0]):
        for j in range(0, mask_image.shape[1]):
            if mask_image[i, j] > 127:
                mask_image[i, j] = 1
            else:
                mask_image[i, j] = 0

    #initialising the four parameters
    TruePositive = 0.001
    TrueNegative = 0.001
    FalsePositive = 0.001
    FalseNegative = 0.001
    for i in range(0, mask_image.shape[0]):
        for j in range(0, mask_image.shape[1]):
            pred = prediction[i, j]
            truth = truth_image[i, j]
            mask = mask_image[i, j]

            if mask == 1:
                #finding the number of matches of each parameter
                if pred == 1 and truth == 1:
                    TruePositive = TruePositive + 1
                elif pred == 0 and truth == 0:
                    TrueNegative = TrueNegative + 1
                elif pred == 0 and truth == 1:
                    FalseNegative = FalseNegative + 1
                elif pred == 1 and truth == 0:
                    FalsePositive = FalsePositive + 1

    #calculating the results(accuracy, sensitivity &amp; specificity)
    accuracy = float((TruePositive + TrueNegative)) / float((TruePositive + FalsePositive + FalseNegative + TrueNegative))
    sensitivity = float((TruePositive)) / float((TruePositive + FalseNegative))
    specificity = float((TrueNegative)) / float((TrueNegative + FalsePositive))

    #to avoid run-time exceptions
    try:
        positivePredictiveValue = float((TruePositive)) / float((TruePositive + FalsePositive))
    except Exception:
        positivePredictiveValue = 0

    #stores the calculated results in an excel sheet
    ws['A' + str(count)] = file
    ws['B' + str(count)] = abs(TruePositive)
    ws['C' + str(count)] = abs(TrueNegative)
    ws['D' + str(count)] = abs(FalsePositive)
    ws['E' + str(count)] = abs(FalseNegative)
    ws['F' + str(count)] = accuracy
    ws['G' + str(count)] = sensitivity
    ws['H' + str(count)] = specificity
    
    count = count + 1

#saving the excel file
wb.save('output.xlsx')